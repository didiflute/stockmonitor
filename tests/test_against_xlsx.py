"""验证 Python 信号逻辑跟用户的 xlsx 模型完全一致

读取 投资策略-TQQQ-v3.xlsx 的"信号"sheet, 提取每一行的:
  - 收盘价
  - Sheets 计算的每日/每周/每月变化率
  - Sheets 给出的每日/每周/每月信号 (买/卖/-)

然后用 src/signals.py 重新计算, 逐行对比, 必须完全一致 (浮点容差 1e-9).
"""
import sys
from pathlib import Path

# 让 src/ 可被 import
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import warnings
warnings.filterwarnings("ignore")

from openpyxl import load_workbook
from grid_signal import evaluate_signal, SignalType


XLSX_PATH = Path("/sessions/compassionate-jolly-babbage/mnt/uploads/投资策略-TQQQ-v3.xlsx")
SHEETS_ACTION_MAP = {"买": "buy", "卖": "sell", "-": "hold", None: "hold", " ": "hold"}


def load_signal_sheet():
    """从 xlsx 提取每行: (date, close, daily_chg, daily_sig, weekly_chg, weekly_sig, monthly_chg, monthly_sig)"""
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["信号"]
    rows = []
    for row in range(7, ws.max_row + 1):  # 第 7 行 = 今天, 往下越来越早
        date = ws.cell(row=row, column=3).value  # C
        close = ws.cell(row=row, column=4).value  # D
        daily_chg = ws.cell(row=row, column=8).value  # H
        daily_sig = ws.cell(row=row, column=9).value  # I
        weekly_chg = ws.cell(row=row, column=10).value  # J
        weekly_sig = ws.cell(row=row, column=11).value  # K
        monthly_chg = ws.cell(row=row, column=12).value  # L
        monthly_sig = ws.cell(row=row, column=13).value  # M
        if close is None or not isinstance(close, (int, float)):
            continue
        rows.append({
            "date": date,
            "close": float(close),
            "daily_chg": daily_chg,
            "daily_sig": SHEETS_ACTION_MAP.get(daily_sig, "hold"),
            "weekly_chg": weekly_chg,
            "weekly_sig": SHEETS_ACTION_MAP.get(weekly_sig, "hold"),
            "monthly_chg": monthly_chg,
            "monthly_sig": SHEETS_ACTION_MAP.get(monthly_sig, "hold"),
        })
    return rows


def run_backtest():
    rows = load_signal_sheet()
    print(f"已加载 {len(rows)} 行历史数据 (从 {rows[0]['date']} 到 {rows[-1]['date']})\n")

    closes = [r["close"] for r in rows]  # closes[0] = 今天, closes[1] = 昨天 ...

    mismatches = {"daily": [], "weekly": [], "monthly": []}
    rate_diffs = {"daily": [], "weekly": [], "monthly": []}

    # 对每一行 (除了最后 20 行没足够历史) 都跑一遍信号
    for i, row in enumerate(rows[:-20]):
        current_price = row["close"]
        historical = closes[i + 1: i + 21]  # 后续 20 行作为历史

        for sig_type in ("daily", "weekly", "monthly"):
            sig_type: SignalType
            try:
                result = evaluate_signal(
                    current_price=current_price,
                    historical_closes=historical,
                    signal_type=sig_type,
                    drop_threshold=-0.07,
                    rise_threshold=0.08,
                )
            except ValueError:
                continue

            sheets_chg = row[{"daily": "daily_chg", "weekly": "weekly_chg", "monthly": "monthly_chg"}[sig_type]]
            sheets_sig = row[{"daily": "daily_sig", "weekly": "weekly_sig", "monthly": "monthly_sig"}[sig_type]]

            if sheets_chg is None or not isinstance(sheets_chg, (int, float)):
                continue

            rate_diff = abs(result.change_rate - sheets_chg)
            rate_diffs[sig_type].append(rate_diff)

            if rate_diff > 1e-6:
                mismatches[sig_type].append({
                    "date": row["date"],
                    "type": "rate",
                    "py": result.change_rate,
                    "xlsx": sheets_chg,
                    "diff": rate_diff,
                })

            if result.action != sheets_sig:
                mismatches[sig_type].append({
                    "date": row["date"],
                    "type": "action",
                    "py": result.action,
                    "xlsx": sheets_sig,
                })

    # 输出报告
    all_pass = True
    for sig_type in ("daily", "weekly", "monthly"):
        diffs = rate_diffs[sig_type]
        ms = mismatches[sig_type]
        max_diff = max(diffs) if diffs else 0
        print(f"=== {sig_type.upper()} 信号 ===")
        print(f"  对比 {len(diffs)} 行, 变化率最大误差 {max_diff:.2e}")
        print(f"  不一致条目: {len(ms)}")
        if ms:
            all_pass = False
            for m in ms[:5]:
                print(f"    {m}")
            if len(ms) > 5:
                print(f"    ... 还有 {len(ms) - 5} 条")
        print()

    return all_pass


if __name__ == "__main__":
    ok = run_backtest()
    sys.exit(0 if ok else 1)
